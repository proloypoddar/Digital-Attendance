import cv2
import face_recognition
import os
import openpyxl
from datetime import datetime
known_faces = {}
pictures_folder = "E:\CODE\PROJECT 1\Digital-Attendance\PIC DATA"
for filename in os.listdir(pictures_folder):
    if filename.endswith(".jpg") or filename.endswith(".png"):
        name = os.path.splitext(filename)[0]
        image_path = os.path.join(pictures_folder, filename)
        known_image = face_recognition.load_image_file(image_path)
        known_image_encoding = face_recognition.face_encodings(known_image)[0]
        known_faces[name] = known_image_encoding
excel_folder = "E:\CODE\PROJECT 1\Digital-Attendance"
excel_file_name = datetime.today().strftime('%d-%m-%Y') + ".xlsx"
excel_file_path = os.path.join(excel_folder, excel_file_name)
if not os.path.exists(excel_folder):
    os.makedirs(excel_folder)
if os.path.exists(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
else:
    workbook = openpyxl.Workbook()
today = datetime.today().strftime('%Y-%m-%d')
if today not in workbook.sheetnames:
    sheet = workbook.create_sheet(title=today)
    sheet['A1'] = 'Name'
    for col_num, name in enumerate(known_faces.keys(), start=2):
        sheet.cell(row=1, column=col_num, value=name)
else:
    sheet = workbook[today]
cap = cv2.VideoCapture(0)
while True:
    ret, frame = cap.read()
    face_locations = face_recognition.face_locations(frame)
    for face_location in face_locations:
        test_encoding = face_recognition.face_encodings(frame, [face_location])[0]
        matches = face_recognition.compare_faces(list(known_faces.values()), test_encoding)
        name = "Unknown"
        if True in matches:
            first_match_index = matches.index(True)
            name = list(known_faces.keys())[first_match_index]
            if name not in [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]:
                sheet.append([name] + ['' for _ in range(len(known_faces))])
            name_row = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)].index(name) + 2
            name_column = list(known_faces.keys()).index(name) + 2
            sheet.cell(row=name_row, column=name_column, value='P')

        top, right, bottom, left = face_location
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

    cv2.imshow('Face Recognition', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

workbook.save(excel_file_path)
cap.release()
cv2.destroyAllWindows()
